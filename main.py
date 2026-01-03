import os
import re
import uuid
import tempfile
import time
import json
import openpyxl
import requests
from datetime import datetime
from urllib.parse import urljoin, urlparse, urlencode, urlsplit, urlunsplit, parse_qs
from lxml import html
from openpyxl.styles import PatternFill
from dotenv import load_dotenv
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from db_driver import ConfigDBDriver
from send_email import send_email_notification
import undetected_chromedriver as uc
from selenium_stealth import stealth
from openpyxl import Workbook
# Load env
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
BASE_URL = "http://86.104.73.3/" #os.getenv("BASE_URL", "http://localhost")

OUTPUT_FOLDER = "output_files"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

GPT_SCHEMA = {
    "Название": None,
    "Цена": None,
    "Валюта": None,
    "Площадь": None,
    "Площадь земли": None,
    "Тип объекта": None,
    "Год постройки": None,
    "Количество комнат": None,
    "Описание": None,
    "Инфраструктура": None,
    "С/у": None,
    "Этаж": None,
    "Локация": None,
    "Координаты": None,
    "Контактное лицо": None,
    "Телефон контактного лица": None,
    "Компания": None,
    "Телефон компании": None
}


# ============================================================
# 🧠 DB Config Loader
# ============================================================
def get_website_config(url):
    db = ConfigDBDriver()
    rows = db.cursor.execute("SELECT website, config_json FROM configs").fetchall()
    for row in rows:
        website = row["website"]
        if website in url:
            db.close()
            return website, json.loads(row["config_json"])
    db.close()
    return None, None


# ============================================================
# 🌐 Selenium Loader
# ============================================================
def get_rendered_html(url, config):
    headless = config.get("headless", True)

    options = uc.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")

    # Set realistic User-Agent
    ua = config.get(
        "user_agent",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"
    )
    options.add_argument(f"--user-agent={ua}")

    # Optional: unique profile per run to avoid session conflicts
    tempdir = tempfile.mkdtemp()
    options.add_argument(f"--user-data-dir={tempdir}")

    driver = None
    try:
        driver = uc.Chrome(options=options)

        # Apply selenium-stealth tweaks
        stealth(driver,
                languages=["en-US", "en"],
                vendor="Google Inc.",
                platform="Win32",
                webgl_vendor="Intel Inc.",
                renderer="Intel Iris OpenGL Engine",
                fix_hairline=True,
                )

        wait_time = config.get("wait_time", 6)
        lazy_scroll = config.get("lazy_scroll", False)
        max_scrolls = config.get("max_scrolls", 10)
        scroll_pause = config.get("scroll_pause", 2)
        page_ready_xpath = config.get("page_ready_xpath", "//*")

        driver.get(url)

        try:
            WebDriverWait(driver, wait_time).until(
                EC.presence_of_all_elements_located((By.XPATH, page_ready_xpath))
            )
        except TimeoutException:
            print("⚠️ Timeout waiting for page load")

        if lazy_scroll:
            last_height = driver.execute_script("return document.body.scrollHeight")
            for _ in range(max_scrolls):
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(scroll_pause)
                new_height = driver.execute_script("return document.body.scrollHeight")
                if new_height == last_height:
                    break
                last_height = new_height

        html_content = driver.page_source
        driver.quit()
        return html_content

    except WebDriverException as e:
        print(f"⚠️ Selenium Error: {e}")
        if driver:
            driver.quit()
        return None
# ============================================================
# 🧩 Property Parser
# ============================================================
# ============================================================
# 🧩 Property Parser (fixed title spacing)
# ============================================================
def gpt_extract_fields(html_content: str, url: str, missing_fields: list):
    system_prompt = "You are a professional real estate data extractor."

    user_prompt = f"""
Extract ONLY the following missing fields from the HTML.

Rules:
- Return RAW JSON
- No explanations
- If value not found → null
- Do NOT invent data

Fields to extract:
{json.dumps(missing_fields, ensure_ascii=False)}

URL:
{url}

HTML:
{html_content[:120000]}
"""

    headers = {
        "Authorization": f"Bearer {os.getenv('CHAT_GPT_API_KEY')}",
        "Content-Type": "application/json"
    }

    payload = {
        "model": os.getenv("CHAT_GPT_MODEL"),
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ],
        "temperature": 0
    }

    r = requests.post(
        os.getenv("CHAT_GPT_URL"),
        headers=headers,
        json=payload,
        timeout=120
    )

    if r.status_code != 200:
        print("⚠️ GPT error:", r.text)
        return {}

    try:
        content = r.json()["choices"][0]["message"]["content"]
        return json.loads(content)
    except Exception as e:
        print("⚠️ GPT JSON error:", e)
        return {}

def parse_property_with_config(url, config, download_folder="images"):
    result = {"Ссылка на объект": url}
    os.makedirs(download_folder, exist_ok=True)

    html_content = get_rendered_html(url, config)
    if not html_content:
        return None, "Failed to load HTML"

    tree = html.fromstring(html_content)
    fields = config.get("fields", {})

    missing_for_gpt = []

    for field_name, field_data in fields.items():
        xpath = field_data.get("xpath")
        transform = field_data.get("transform")

        if not xpath:
            result[field_name] = "ERROR"
            missing_for_gpt.append(field_name)
            continue

        values = tree.xpath(xpath)
        if not values:
            result[field_name] = "ERROR"
            missing_for_gpt.append(field_name)
            continue

        cleaned = []
        for v in values:
            txt = v.text_content().strip() if hasattr(v, "text_content") else str(v).strip()
            if txt:
                cleaned.append(txt)

        combined = "\n".join(dict.fromkeys(cleaned))  # remove duplicates

        if transform:
            try:
                combined = eval(transform, {"re": re}, {"value": combined})
            except:
                combined = combined

        if not combined:
            result[field_name] = "ERROR"
            missing_for_gpt.append(field_name)
        else:
            result[field_name] = combined

    # 🤖 GPT FALLBACK
    if missing_for_gpt:
        print(f"🧠 GPT extracting missing fields: {missing_for_gpt}")
        gpt_data = gpt_extract_fields(html_content, url, missing_for_gpt)

        for k in missing_for_gpt:
            if k in gpt_data and gpt_data[k]:
                result[k] = gpt_data[k]
            elif result.get(k) == "ERROR":
                result[k] = "ERROR"

    return result, None

# ============================================================
# 💾 Excel Export (fixed column order in Russian)
# ============================================================


def save_to_excel(properties, filename, output_folder="output_files"):
    os.makedirs(output_folder, exist_ok=True)

    # ✅ Desired column order (Russian)
    desired_order = [
        "Ссылка на объект", "Название", "Цена", "Валюта", "Площадь", "Площадь земли",
        "Тип объекта", "Год постройки", "Количество комнат", "Описание", "Инфраструктура",
        "С/у", "Этаж", "Локация", "Координаты", "Фото_ссылки", "Фото_уникальные_названия",
        "Контактное лицо", "Телефон контактного лица", "Компания", "Телефон компании"
    ]

    # Ensure all properties have all columns
    for prop in properties:
        for col in desired_order:
            if col not in prop:
                prop[col] = "ERROR"

    wb = Workbook()
    ws = wb.active
    ws.append(desired_order)

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    for prop in properties:
        row = []
        for col in desired_order:
            value = prop.get(col, "ERROR")

            # ✅ If value is a list, join it into a string
            if isinstance(value, list):
                value = ";".join([str(v) for v in value])

            # ✅ If value is None, replace with empty string
            if value is None:
                value = ""

            row.append(value)

        ws.append(row)

        # Fill red for ERROR cells
        for i, value in enumerate(row, start=1):
            if value == "ERROR":
                ws.cell(row=ws.max_row, column=i).fill = red_fill

    file_path = os.path.join(output_folder, filename)
    wb.save(file_path)
    return file_path

# ============================================================
# 🧩 List Page Parser with Auto Pagination (Next Button Supported)
# ============================================================
def parse_list_page(base_url, config):
    print(f"🌍 Fetching list pages from: {base_url}")

    properties = []
    seen_first = None
    page = 1
    page_query = config.get("page_query")
    next_button_xpath = config.get("next_page_xpath")

    headless = config.get("headless", True)

    options = uc.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    ua = config.get(
        "user_agent",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"
    )
    options.add_argument(f"--user-agent={ua}")
    tempdir = tempfile.mkdtemp()
    options.add_argument(f"--user-data-dir={tempdir}")

    driver = uc.Chrome(options=options)

    # Apply stealth
    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
            )

    driver.get(base_url)
    time.sleep(3)

    while True:
        print(f"\n🔄 Loading page {page}...")

        html_content = driver.page_source
        tree = html.fromstring(html_content)
        property_links = tree.xpath(config.get("list_page_check", ""))

        if not property_links:
            print("🚫 No property links found → stopping pagination.")
            break

        first_url = urljoin(base_url, property_links[0])
        if seen_first == first_url:
            print("🛑 Same first record as previous → last page reached.")
            break
        seen_first = first_url

        for idx, link in enumerate(property_links, start=1):
            full_url = urljoin(base_url, link)
            print(f"➡️ [{idx}/{len(property_links)}] {full_url}")
            data, error = parse_property_with_config(full_url, config)
            if data:
                properties.append(data)
            else:
                print(f"❌ Error parsing property: {error}")

        # Pagination
        if next_button_xpath:
            try:
                next_btn = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, next_button_xpath))
                )
                driver.execute_script("arguments[0].scrollIntoView(true);", next_btn)
                time.sleep(1)
                driver.execute_script("arguments[0].click();", next_btn)
                print("👉 Clicked next page button via JS.")
                time.sleep(config.get("scroll_pause", 2))
                page += 1
                continue
            except Exception as e:
                print(f"🛑 No next page button found or not clickable: {e}")
                break
        elif page_query:
            parts = list(urlsplit(base_url))
            query = parse_qs(parts[3])
            query[page_query] = [str(page + 1)]
            parts[3] = urlencode(query, doseq=True)
            next_page_url = urlunsplit(parts)
            print(f"➡️ Loading next page via query: {next_page_url}")
            driver.get(next_page_url)
            time.sleep(2)
            page += 1
            continue
        else:
            break

    driver.quit()
    return properties
# ============================================================
# 🤖 Telegram Bot Handlers
# ============================================================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("👋 Отправьте URL объекта недвижимости/листинга. Я скачаю и экспортирую всё в Excel.")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    url = update.message.text.strip()
    await update.message.reply_text("🔍 Собираю данные…")

    domain, config = get_website_config(url)
    if not config:
        await update.message.reply_text("❌ Источник не подключён.")
        return

    data, error = parse_property_with_config(url, config)

    if data and data.get("Название") != "ERROR":
        properties = [data]
    else:
        properties = parse_list_page(url, config)

    if not properties:
        await update.message.reply_text("❌ Ничего не найдено.")
        return

    filename = f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}_properties.xlsx"
    save_to_excel(properties, filename)

    await update.message.reply_text(
        f"✅ Готово: {len(properties)} объектов\n📂 {BASE_URL}/output_files/{filename}"
    )

# ============================================================
# 🚀 Launch Bot
# ============================================================
if __name__ == "__main__":
    print("🤖 Bot running — config-driven, paginated scraper active...")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.run_polling()
